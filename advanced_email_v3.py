import asyncio
import re
import json
import csv
import time
import random
import aiohttp
import os
from urllib.parse import urljoin, urlparse
from playwright.async_api import async_playwright
from collections import deque
from pathlib import Path
import logging
from typing import Set, Dict, List
import subprocess

# =============== DEPENDENCIES ===============
# pip install playwright asyncio aiohttp openpyxl python-docx PyPDF2
# playwright install chromium
# sudo apt install tor poppler-utils antiword

# =============== TACTICAL CONFIG ===============
MAX_PAGES = 1000
MAX_SUBDOMAINS = 30
CONCURRENT = 25
DELAY = (0.3, 1.2)
TIMEOUT = 8
OUTPUT_DIR = "recon_reports"

# AI Keywords for role classification
SENIORITY_KEYWORDS = {
    "executive": ["director", "head", "dean", "chair", "provost", "vp", "vice president"],
    "senior": ["professor", "principal", "lead", "chief", "manager", "coordinator"],
    "mid": ["lecturer", "engineer", "specialist", "officer", "analyst"],
    "junior": ["assistant", "trainee", "intern", "student", "researcher"]
}

DEPT_KEYWORDS = {
    "Computer Science": ["cs", "computer", "software", "ai", "data", "it"],
    "Electrical Engineering": ["eee", "electrical", "electronics", "power", "telecom"],
    "Mechanical": ["me", "mechanical", "thermal", "automotive", "robotics"],
    "Civil": ["ce", "civil", "construction", "structural", "environmental"],
    "Admin": ["admission", "registrar", "finance", "hr", "library", "accounts"]
}

# =============== KING INTELLIGENCE CORE ===============
class KingHarvester:
    def __init__(self, root_url: str):
        self.root_url = root_url.rstrip('/')
        self.base_domain = urlparse(root_url).netloc.lower()
        if self.base_domain.startswith('www.'):
            self.base_domain = self.base_domain[4:]
        
        # Intelligence database
        self.profiles: Dict[str, Dict] = {}  # email -> full intel
        self.visited: Set[str] = set()
        self.documents: Set[str] = set()
        self.pages = 0
        self.session = None
        self.browser = None
        
        # Create output dir
        self.out_dir = Path(OUTPUT_DIR) / f"{self.base_domain}_{int(time.time())}"
        self.out_dir.mkdir(parents=True, exist_ok=True)

    # ==================== AI CLASSIFIER ====================
    def classify_profile(self, text: str, email: str) -> Dict:
        """AI-powered role/seniority/department prediction"""
        # Clean text
        clean_text = re.sub(r'\s+', ' ', text.lower())
        
        # Name extraction (heuristic)
        name = ""
        email_local = email.split('@')[0]
        name_parts = re.split(r'[._-]', email_local)
        if len(name_parts) >= 2 and all(len(p) > 1 for p in name_parts[:2]):
            name = " ".join(p.capitalize() for p in name_parts[:2])
        
        # Seniority classification
        seniority = "unknown"
        max_score = 0
        for level, keywords in SENIORITY_KEYWORDS.items():
            score = sum(1 for kw in keywords if kw in clean_text)
            if score > max_score:
                max_score = score
                seniority = level
        
        # Department classification
        department = "General"
        dept_score = 0
        for dept, keywords in DEPT_KEYWORDS.items():
            score = sum(1 for kw in keywords if kw in clean_text)
            if score > dept_score:
                dept_score = score
                department = dept
        
        # Confidence score (0.0 - 1.0)
        confidence = 0.3  # baseline
        if name:
            confidence += 0.2
        if seniority != "unknown":
            confidence += 0.15
        if department != "General":
            confidence += 0.15
        if max_score + dept_score > 2:
            confidence += 0.2
        
        return {
            "name": name,
            "seniority": seniority,
            "department": department,
            "confidence": min(confidence, 0.95),
            "raw_snippet": self._get_snippet(text, email)
        }

    def _get_snippet(self, text: str, email: str) -> str:
        """Extract 200-char context around email"""
        lines = text.split('\n')
        for line in lines:
            if email in line.lower():
                start = max(0, line.lower().index(email) - 100)
                return line[start:start+200].strip()
        return "Context not found"

    # ==================== DEEP DISCOVERY ENGINE ====================
    async def discover_subdomains(self) -> List[str]:
        """Passive subdomain discovery (crt.sh + common)"""
        subdomains = set()
        try:
            url = f"https://crt.sh/?q=%25.{self.base_domain}&output=json"
            async with self.session.get(url) as resp:
                if resp.status == 200:
                    data = await resp.json(content_type=None)
                    for entry in data:
                        name = entry.get('name_value', '').lower().strip()
                        if name and name.endswith(self.base_domain) and '*' not in name:
                            subdomains.add(f"https://{name}")
        except:
            pass
        
        # Add tactical subdomains
        tactical_subs = [
            "www", "mail", "webmail", "staff", "faculty", "cs", "it", "eee", "me", "ce",
            "admissions", "students", "library", "research", "alumni", "contact",
            "portal", "intranet", "admin", "accounts", "hr", "finance", "registrar"
        ]
        for sub in tactical_subs:
            subdomains.add(f"https://{sub}.{self.base_domain}")
        
        return list(subdomains)[:MAX_SUBDOMAINS]

    # ==================== DOCUMENT MINER ====================
    async def download_document(self, url: str) -> str:
        """Download and save document"""
        try:
            async with self.session.get(url) as resp:
                if resp.status == 200:
                    ext = Path(urlparse(url).path).suffix.lower()
                    if ext in {'.pdf', '.doc', '.docx', '.ppt', '.pptx'}:
                        filename = self.out_dir / f"doc_{abs(hash(url))}{ext}"
                        with open(filename, 'wb') as f:
                            f.write(await resp.read())
                        return str(filename)
        except:
            pass
        return ""

    def extract_emails_from_doc(self, filepath: str) -> Set[str]:
        """Extract emails from documents"""
        emails = set()
        try:
            text = ""
            if filepath.endswith('.pdf'):
                result = subprocess.run(['pdftotext', '-layout', filepath, '-'], 
                                      capture_output=True, text=True)
                text = result.stdout
            else:
                result = subprocess.run(['strings', filepath], capture_output=True, text=True)
                text = result.stdout
            
            std_emails = re.findall(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text)
            for email in std_emails:
                if self.is_target_domain(email):
                    emails.add(email.lower())
        except:
            pass
        return emails

    # ==================== STEALTH BROWSER ====================
    async def init_browser(self):
        pw = await async_playwright().start()
        # Optional: Add proxy support
        # launch_args = {"proxy": {"server": "socks5://127.0.0.1:9150"}}  # Tor
        self.browser = await pw.chromium.launch(
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-web-security",
                "--disable-features=IsolateOrigins",
                "--disable-extensions",
                "--disable-plugins"
            ]
        )

    async def fetch_with_evasion(self, url: str) -> str:
        if not self.browser:
            await self.init_browser()
        
        for attempt in range(2):
            context = await self.browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                viewport={"width": 1920, "height": 1080},
                locale="en-US",
                timezone_id="America/New_York"
            )
            await context.add_init_script("""
                delete navigator.__proto__.webdriver;
                window.chrome = {runtime: {}};
                Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
            """)
            
            page = await context.new_page()
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=TIMEOUT*1000)
                await page.wait_for_timeout(random.uniform(1000, 2500))
                return await page.content()
            except:
                pass
            finally:
                await context.close()
            await asyncio.sleep(2)
        return ""

    # ==================== CORE INTELLIGENCE GATHERING ====================
    def is_target_domain(self, email: str) -> bool:
        if '@' not in email:
            return False
        _, domain = email.lower().rsplit('@', 1)
        return domain == self.base_domain or domain.endswith('.' + self.base_domain)

    def extract_and_clean_emails(self, text: str) -> Set[str]:
        emails = set()
        # Standard emails
        std_emails = re.findall(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text)
        for email in std_emails:
            if self.is_target_domain(email):
                emails.add(email.lower())
        # Add de-obfuscation if needed
        return emails

    async def scrape_page(self, url: str):
        if url in self.visited or self.pages >= MAX_PAGES:
            return []
        self.visited.add(url)
        self.pages += 1
        print(f"[{self.pages:4d}] 🕵️ {url[:60]}...")

        html = await self.fetch_with_evasion(url)
        if not html:
            return []

        # Extract emails + AI classification
        new_emails = self.extract_and_clean_emails(html)
        for email in new_emails:
            if email not in self.profiles:
                ai_data = self.classify_profile(html, email)
                self.profiles[email] = {
                    "email": email,
                    "name": ai_data["name"],
                    "seniority": ai_data["seniority"],
                    "department": ai_data["department"],
                    "confidence": round(ai_data["confidence"], 2),
                    "source_url": url,
                    "context_snippet": ai_data["raw_snippet"]
                }
                # Create individual dossier
                self._write_dossier(email, self.profiles[email])

        # Mine documents
        doc_links = re.findall(r'href=[\'"]?([^\'" >]+\.(pdf|docx?|pptx?))', html, re.IGNORECASE)
        for link, _ in doc_links:
            full_url = urljoin(url, link)
            if full_url not in self.documents:
                self.documents.add(full_url)
                doc_path = await self.download_document(full_url)
                if doc_path:
                    doc_emails = self.extract_emails_from_doc(doc_path)
                    for email in doc_emails:
                        if email not in self.profiles:
                            self.profiles[email] = {
                                "email": email,
                                "name": "",
                                "seniority": "unknown",
                                "department": "General",
                                "confidence": 0.4,
                                "source_url": full_url,
                                "context_snippet": "From document"
                            }
                            self._write_dossier(email, self.profiles[email])

        # Next links
        next_urls = []
        if self.pages < MAX_PAGES:
            links = re.findall(r'href=[\'"]?([^\'" >]+)', html)
            for raw in links:
                if raw.startswith(('http://', 'https://')):
                    full = raw
                elif raw.startswith('/'):
                    full = urljoin(url, raw)
                else:
                    continue
                parsed = urlparse(full)
                domain = parsed.netloc.lower()
                if domain == self.base_domain or domain.endswith('.' + self.base_domain):
                    if full not in self.visited:
                        next_urls.append(full)
        return next_urls

    def _write_dossier(self, email: str, data: Dict):
        """Write human-readable dossier file"""
        safe_email = email.replace('@', '_at_').replace('.', '_')
        dossier_path = self.out_dir / f"dossier_{safe_email}.txt"
        with open(dossier_path, 'w', encoding='utf-8') as f:
            f.write(f"📧 EMAIL INTELLIGENCE DOSSIER\n")
            f.write(f"{'='*50}\n")
            f.write(f"Target Email     : {data['email']}\n")
            f.write(f"Name             : {data['name'] or 'Not identified'}\n")
            f.write(f"Seniority Level  : {data['seniority'].title()}\n")
            f.write(f"Department       : {data['department']}\n")
            f.write(f"Confidence Score : {data['confidence']:.0%}\n")
            f.write(f"Source URL       : {data['source_url']}\n")
            f.write(f"Context Snippet  : {data['context_snippet']}\n")
            f.write(f"\n[AI Analysis]\n")
            if data['confidence'] >= 0.8:
                f.write("✅ HIGH CONFIDENCE: Likely accurate profile\n")
            elif data['confidence'] >= 0.6:
                f.write("⚠️  MEDIUM CONFIDENCE: Verify manually\n")
            else:
                f.write("❌ LOW CONFIDENCE: Treat as unverified\n")

    # ==================== EXCEL MASTER OUTPUT ====================
    def export_to_excel(self):
        """Create professional Excel report"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("⚠️  openpyxl not installed. Installing...")
            os.system('pip install openpyxl')
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Email Intelligence"

        # Header
        headers = ["Email", "Name", "Seniority", "Department", "Confidence", "Source URL", "Context Snippet"]
        ws.append(headers)
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col in range(1, len(headers)+1):
            ws.cell(1, col).font = header_font
            ws.cell(1, col).fill = header_fill
            ws.cell(1, col).alignment = Alignment(horizontal="center")

        # Data rows
        for email, data in self.profiles.items():
            row = [
                data['email'],
                data['name'],
                data['seniority'].title(),
                data['department'],
                f"{data['confidence']:.0%}",
                data['source_url'],
                data['context_snippet'][:100] + "..." if len(data['context_snippet']) > 100 else data['context_snippet']
            ]
            ws.append(row)
            
            # Color-code by confidence
            confidence = data['confidence']
            if confidence >= 0.8:
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
            elif confidence >= 0.6:
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
            else:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
            
            for col in range(1, len(headers)+1):
                ws.cell(ws.max_row, col).fill = fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        excel_path = self.out_dir / f"EMAIL_INTELLIGENCE_{self.base_domain.upper()}.xlsx"
        wb.save(excel_path)
        return excel_path

    # ==================== MAIN CRAWL LOOP ====================
    async def init_session(self):
        self.session = aiohttp.ClientSession(
            timeout=aiohttp.ClientTimeout(total=TIMEOUT),
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        )

    async def crawl(self):
        await self.init_session()
        
        # Start URLs
        start_urls = [self.root_url]
        subdomains = await self.discover_subdomains()
        start_urls.extend(subdomains)
        sensitive_paths = ["/staff", "/faculty", "/people", "/directory", "/contact", "/about", "/team"]
        for path in sensitive_paths:
            start_urls.append(urljoin(self.root_url, path))
        
        queue = deque(start_urls)
        tasks = set()
        
        while (queue or tasks) and self.pages < MAX_PAGES:
            while queue and len(tasks) < CONCURRENT:
                url = queue.popleft()
                task = asyncio.create_task(self._worker(url, queue))
                tasks.add(task)
                task.add_done_callback(tasks.discard)
            await asyncio.sleep(0.01)
        
        print(f"\n👑 KING RECON COMPLETE: {len(self.profiles)} profiles from {self.pages} assets")

    async def _worker(self, url: str, queue: deque):
        next_urls = await self.scrape_page(url)
        for u in next_urls:
            queue.append(u)
        await asyncio.sleep(random.uniform(*DELAY))

    async def close(self):
        if self.session:
            await self.session.close()
        if self.browser:
            await self.browser.close()

    def finalize(self):
        """Generate final reports"""
        excel_path = self.export_to_excel()
        summary_path = self.out_dir / "RECON_SUMMARY.txt"
        
        with open(summary_path, 'w') as f:
            f.write(f"🎯 KING HARVESTER RECON REPORT\n")
            f.write(f"{'='*40}\n")
            f.write(f"Target Domain: {self.base_domain}\n")
            f.write(f"Total Profiles: {len(self.profiles)}\n")
            f.write(f"Pages Crawled: {self.pages}\n")
            f.write(f"Documents Analyzed: {len(self.documents)}\n\n")
            
            # Confidence breakdown
            high = len([p for p in self.profiles.values() if p['confidence'] >= 0.8])
            medium = len([p for p in self.profiles.values() if 0.6 <= p['confidence'] < 0.8])
            low = len([p for p in self.profiles.values() if p['confidence'] < 0.6])
            
            f.write(f"Confidence Breakdown:\n")
            f.write(f"  HIGH (≥80%)   : {high}\n")
            f.write(f"  MEDIUM (60-79%): {medium}\n")
            f.write(f"  LOW (<60%)     : {low}\n\n")
            
            f.write(f"📁 Output Directory: {self.out_dir.absolute()}\n")
            f.write(f"  - Excel Master Report: {excel_path.name}\n")
            f.write(f"  - Individual Dossiers: dossier_*.txt\n")
            f.write(f"  - Full Recon Summary: {summary_path.name}\n")
        
        print(f"\n✅ KING REPORT GENERATED:\n   {summary_path}")
        print(f"📊 EXCEL MASTER: {excel_path}")

# =============== EXECUTE ===============
async def main():
    target = input("👑 Enter target domain (e.g., duet.edu.pk): ").strip()
    if not target.startswith(("http://", "https://")):
        target = "https://" + target
    
    print(f"\n🚀 LAUNCHING KING HARVESTER v4.0 (MILITARY EDITION)")
    print(f"   Target: {target}")
    print(f"   Max Pages: {MAX_PAGES}")
    print(f"   Output: {OUTPUT_DIR}/")
    
    harvester = KingHarvester(target)
    try:
        await harvester.crawl()
        harvester.finalize()
    finally:
        await harvester.close()

if __name__ == "__main__":
    asyncio.run(main())
